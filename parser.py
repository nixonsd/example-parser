import csv
import urllib.request
from bs4 import BeautifulSoup
from random import choice, uniform
import openpyxl


def get_html(url, useragent=None, proxy=None):
    proxy_support = urllib.request.ProxyHandler(proxy)
    opener = urllib.request.build_opener(proxy_support)
    urllib.request.install_opener(opener)
    req = urllib.request.Request(url, headers=useragent)
    response = urllib.request.urlopen(req)
    return response.read()

def parse(html):
    global pages
    soup = BeautifulSoup(html, features="lxml")
    pages = soup.find_all(attrs={'class': "page-numbers"})[-2].string
    table = soup.find("div", id="content")
    return table

def main():

    useragents = open('useragents.txt').read().split('\n')
    proxies = open('proxies.txt').read().split('\n')

    url = "https://fobook.ru/page/1/"
    while True:
        try:
            html = get_html(url, useragent={'User-Agent':"{agent}".format(agent=choice(useragents))}, proxy={'https': 'https://' + choice(proxies)})
            break
        except:
            continue

    #Код отвечающий за парс
    table = parse(html)

    wb = openpyxl.load_workbook(filename = 'output.xlsx')
    sheet = wb.active
    sheet['A1'] = 'Books'
    sheet['B1'] = 'Level'
    sheet['C1'] = 'Author'

    j = 2
    for i in range(1, int(pages) + 1):
        print("Страница №: {page}".format(page=i))
        url = "https://fobook.ru/page/{page}/".format(page=i)
        while True:
            try:
                html = get_html(url, useragent={'User-Agent':"{agent}".format(agent=choice(useragents))}, proxy={'https': 'https://' + choice(proxies)})
                break
            except:
                continue
        table = parse(html)
        for item in table.find_all("article"):
            print(item.find_all('a')[0].string, "-", item.find_all('a')[1].string, '( Author:', item.find("h3", class_="author").string, ')')
            sheet.cell(row=j, column = 1).value = item.find_all('a')[0].string
            sheet.cell(row=j, column = 2).value = item.find_all('a')[1].string
            sheet.cell(row=j, column = 3).value = item.find("h3", class_="author").string
            j += 1

        print()

    wb.save('output.xlsx')

if __name__ == '__main__':
    main()